import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.core.errors import ExcelNotUploadedError
from app.repositories.excel_repository import ExcelRepository
from app.schemas.domain import RecommendationItem
from app.schemas.requests import CreateFormRequest, FromRecommendationRequest
from app.services.form_service import FormService
from app.services.room_catalog_service import RoomCatalogService


class ExcelAndFormTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.excel_path = Path(self.tmpdir.name) / "hotel.xlsx"
        ExcelRepository.create_template(self.excel_path)

    def tearDown(self):
        self.tmpdir.cleanup()

    def test_append_and_update_pending_form(self):
        catalog = RoomCatalogService()
        repository = ExcelRepository(self.excel_path)
        service = FormService(repository, catalog)
        created = service.create_form(
            CreateFormRequest(
                contact_name="张三",
                gender="男",
                phone="13800000000",
                total_amount=268,
                guest_count=2,
                guest_type="多人",
                stay_days=1,
                room_numbers=["202"],
                order_status="待完成",
            )
        )
        self.assertEqual(catalog.get_room("202").status, "已预订")
        pending = service.list_pending_forms()
        self.assertEqual(len(pending), 1)

        completed = service.complete_pending_form(created.form_id)
        self.assertEqual(completed.order_status, "已完成")
        self.assertEqual(catalog.get_room("202").status, "已住")
        self.assertEqual(service.list_pending_forms(), [])

    def test_create_completed_form_sets_room_occupied(self):
        catalog = RoomCatalogService()
        repository = ExcelRepository(self.excel_path)
        service = FormService(repository, catalog)
        created = service.create_form(
            CreateFormRequest(
                contact_name="李四",
                gender="女",
                phone="13900000000",
                total_amount=188,
                guest_count=1,
                guest_type="个人",
                stay_days=1,
                room_numbers=["201"],
                order_status="已完成",
            )
        )
        self.assertEqual(created.room_status_result, "已住")
        self.assertEqual(catalog.get_room("201").status, "已住")

    def test_create_from_recommendation_allows_manual_room_selection(self):
        catalog = RoomCatalogService()
        repository = ExcelRepository(self.excel_path)
        service = FormService(repository, catalog)
        recommendation = RecommendationItem(
            recommendation_id="rec_001",
            room_numbers=["202", "203"],
            selectable_room_numbers=["202", "203", "205", "301"],
            room_signature="多人间:普通房::268.00|多人间:普通房::268.00",
            total_amount=1608,
            guest_count=8,
            room_count=2,
            stay_days=3,
            guest_type="多人",
            reason_text="demo",
            reply_text="demo",
        )
        form = service.create_from_recommendation(
            FromRecommendationRequest(
                session_id="s1",
                recommendation_id="rec_001",
                contact_name="王五",
                gender="男",
                phone="13800000000",
                selected_room_numbers=["205", "301"],
                order_status="待完成",
            ),
            recommendation,
        )
        self.assertEqual(form.room_numbers, ["205", "301"])
        self.assertEqual(catalog.get_room("205").status, "已预订")
        self.assertEqual(catalog.get_room("301").status, "已预订")

    def test_unuploaded_excel_blocks_form_creation(self):
        service = FormService(None, RoomCatalogService())
        with self.assertRaises(ExcelNotUploadedError):
            service.create_form(
                CreateFormRequest(
                    contact_name="张三",
                    gender="男",
                    phone="13800000000",
                    total_amount=188,
                    guest_count=1,
                    guest_type="个人",
                    stay_days=1,
                    room_numbers=["201"],
                )
            )

    def test_missing_header_is_initialized(self):
        workbook = load_workbook(self.excel_path)
        sheet = workbook["订单表单"]
        sheet.cell(row=1, column=2, value="错误表头")
        workbook.save(self.excel_path)
        workbook.close()

        repository = ExcelRepository(self.excel_path)
        repository.initialize_template()

        workbook = load_workbook(self.excel_path)
        sheet = workbook["订单表单"]
        self.assertEqual(sheet.cell(row=1, column=2).value, "联系人")
        workbook.close()

    def test_legacy_form_sheet_name_is_accepted(self):
        workbook = load_workbook(self.excel_path)
        workbook["订单表单"].title = "订单表格"
        workbook.save(self.excel_path)
        workbook.close()

        repository = ExcelRepository(self.excel_path)
        repository.validate()


if __name__ == "__main__":
    unittest.main()
