from pathlib import Path
import shutil
import tempfile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.errors import ExcelNotUploadedError, ExcelRepositoryError, NotFoundError
from app.schemas.domain import OrderForm

SHEET_FORMS = "订单表单"
SHEET_ROOM_STATUS = "房态"
FORM_SHEET_ALIASES = (SHEET_FORMS, "订单工作表", "订单表格")

FORM_HEADERS = [
    "表单唯一标识",
    "联系人",
    "性别",
    "手机号",
    "总金额",
    "人数",
    "人数类型",
    "入住日期",
    "居住天数",
    "房间号",
    "订单状态",
    "房态同步结果",
    "创建时间",
    "更新时间",
]

ROOM_STATUS_HEADERS = ["房间号", "房态", "更新时间"]


class ExcelRepository:
    def __init__(self, file_path: Path | str | None) -> None:
        self.file_path = Path(file_path) if file_path else None

    @staticmethod
    def create_template(file_path: Path | str) -> Path:
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        forms = workbook.active
        forms.title = SHEET_FORMS
        forms.append(FORM_HEADERS)
        room_status = workbook.create_sheet(SHEET_ROOM_STATUS)
        room_status.append(ROOM_STATUS_HEADERS)
        workbook.save(path)
        return path

    def _ensure_uploaded(self) -> Path:
        if self.file_path is None:
            raise ExcelNotUploadedError()
        if not self.file_path.exists():
            raise ExcelRepositoryError("EXCEL_FILE_NOT_FOUND", f"Excel 文件不存在：{self.file_path}")
        return self.file_path

    def _load(self):
        path = self._ensure_uploaded()
        try:
            workbook = load_workbook(path)
        except PermissionError as exc:
            raise ExcelRepositoryError("EXCEL_WRITE_FAILED", "Excel 文件被占用，无法写入。请关闭文件后重试。") from exc
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ExcelRepositoryError("EXCEL_INVALID_FORMAT", "Excel 文件格式错误，无法读取。请上传有效的 .xlsx 文件。") from exc
        return workbook

    def initialize_template(self) -> None:
        workbook = self._load()
        self._ensure_workbook_structure(workbook)
        self._atomic_save(workbook)

    def validate(self) -> None:
        workbook = self._load()
        self._ensure_workbook_structure(workbook)
        workbook.close()

    def append_form(self, form: OrderForm) -> None:
        workbook = self._load()
        sheet = self._get_form_sheet(workbook)
        if self._find_form_row(sheet, form.form_id):
            workbook.close()
            raise ExcelRepositoryError("FORM_DUPLICATED", f"表单 {form.form_id} 已存在，不能重复新增。")
        sheet.append(self._form_to_row(form))
        self._atomic_save(workbook)

    def update_form(self, form: OrderForm) -> None:
        workbook = self._load()
        sheet = self._get_form_sheet(workbook)
        row_index = self._find_form_row(sheet, form.form_id)
        if not row_index:
            workbook.close()
            raise NotFoundError("FORM_NOT_FOUND", f"表单 {form.form_id} 不存在，无法更新。")
        for column_index, value in enumerate(self._form_to_row(form), start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
        self._atomic_save(workbook)

    def list_pending_forms(self) -> list[dict[str, str]]:
        workbook = self._load()
        sheet = self._get_form_sheet(workbook)
        rows = []
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            if record.get("订单状态") == "待完成":
                rows.append(record)
        workbook.close()
        return rows

    def get_form_record(self, form_id: str) -> dict[str, str] | None:
        workbook = self._load()
        sheet = self._get_form_sheet(workbook)
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            if record.get("表单唯一标识") == form_id:
                workbook.close()
                return record
        workbook.close()
        return None

    def update_room_status(self, room_number: str, status: str) -> None:
        workbook = self._load()
        sheet = workbook[SHEET_ROOM_STATUS]
        now_text = __import__("datetime").datetime.now().isoformat(timespec="seconds")
        target_row = None
        for row_index in range(2, sheet.max_row + 1):
            if str(sheet.cell(row=row_index, column=1).value) == room_number:
                target_row = row_index
                break
        if target_row is None:
            sheet.append([room_number, status, now_text])
        else:
            sheet.cell(row=target_row, column=2, value=status)
            sheet.cell(row=target_row, column=3, value=now_text)
        self._atomic_save(workbook)

    def _atomic_save(self, workbook) -> None:
        path = self._ensure_uploaded()
        tmp_name = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp_name = tmp.name
            workbook.save(tmp_name)
            workbook.close()
            shutil.move(tmp_name, path)
        except PermissionError as exc:
            raise ExcelRepositoryError("EXCEL_WRITE_FAILED", "Excel 文件被占用，无法写入。请关闭文件后重试。") from exc
        except OSError as exc:
            raise ExcelRepositoryError("EXCEL_WRITE_FAILED", f"Excel 写入失败：{exc}") from exc
        finally:
            if tmp_name and Path(tmp_name).exists():
                Path(tmp_name).unlink(missing_ok=True)

    @staticmethod
    def _find_form_row(sheet, form_id: str) -> int | None:
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_index, column=1).value == form_id:
                return row_index
        return None

    @staticmethod
    def _get_form_sheet(workbook):
        for sheet_name in FORM_SHEET_ALIASES:
            if sheet_name in workbook.sheetnames:
                return workbook[sheet_name]
        if len(workbook.sheetnames) == 1:
            sheet = workbook[workbook.sheetnames[0]]
            sheet.title = SHEET_FORMS
            return sheet
        return workbook.create_sheet(SHEET_FORMS, 0)

    @staticmethod
    def _ensure_sheet_headers(sheet, headers: list[str]) -> bool:
        changed = False
        for column_index, header in enumerate(headers, start=1):
            if sheet.cell(row=1, column=column_index).value != header:
                sheet.cell(row=1, column=column_index, value=header)
                changed = True
        return changed

    def _ensure_workbook_structure(self, workbook) -> bool:
        changed = False
        form_sheet = self._get_form_sheet(workbook)
        if form_sheet.title != SHEET_FORMS:
            form_sheet.title = SHEET_FORMS
            changed = True
        changed = self._ensure_sheet_headers(form_sheet, FORM_HEADERS) or changed
        if SHEET_ROOM_STATUS in workbook.sheetnames:
            room_status_sheet = workbook[SHEET_ROOM_STATUS]
        else:
            room_status_sheet = workbook.create_sheet(SHEET_ROOM_STATUS)
            changed = True
        changed = self._ensure_sheet_headers(room_status_sheet, ROOM_STATUS_HEADERS) or changed
        return changed

    @staticmethod
    def _form_to_row(form: OrderForm) -> list[str | int | float]:
        return [
            form.form_id,
            form.contact_name,
            form.gender,
            form.phone,
            form.total_amount,
            form.guest_count,
            form.guest_type,
            form.checkin_date.isoformat(),
            form.stay_days,
            ",".join(form.room_numbers),
            form.order_status,
            form.room_status_result,
            form.created_at.isoformat(timespec="seconds"),
            form.updated_at.isoformat(timespec="seconds"),
        ]
